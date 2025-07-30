import visa
import time
from openpyxl import Workbook

# 配置 

# Excel 文件配置
EXCEL_FILE_PATH = r'E:\AutomaticEfficiencyForPlusCurrent\test_data_ac_extended.xlsx'
VIN_COL = 'A'
IIN_COL = 'B'
VOUT_COL = 'C'
IOUT_COL = 'D'
EFF_SYS_COL = 'E'
PLOSS_SYS_COL = 'F'

# 测量配置
IIN_SHUNT_RESISTOR = 0.01  # 输入电流采样电阻
IOUT_SHUNT_RESISTOR = 0.001  # 输出电流采样电阻

# 设备 VISA 地址
LOAD_ADDRESS     = 'ASRL7::INSTR'
VIN_ADDRESS      = 'USB0::0x1698::0x0837::002000002608::INSTR'
VCC_ADDRESS      = 'USB0::0x1AB1::0x0E11::DP8C193003485::INSTR'
DAQ_ADDRESS      = 'USB0::0x0957::0x2007::MY49029470::INSTR'

# 交流模式额外配置
AC_MODE          = False           # 设置为 True 启用交流模式
AC_DUTY_CYCLE    = 50.0            # 占空比 (%)
AC_WAVEFORM      = 'SINE'          # 波形类型: SINE, SQUARE, TRIANGLE
AC_FREQUENCY     = 50.0            # 交流频率 (Hz)

# 初始化 Excel

try:
    efficiency_file = Workbook()
    print('成功创建新的 Excel 工作簿。')
except IOError:
    print('错误：创建 Excel 工作簿失败。')
    exit()

ws = efficiency_file.active
ws.title = "效率数据"
# 写入表头
ws[f'{VIN_COL}1'] = '输入电压 (V)'
ws[f'{IIN_COL}1'] = '输入电流 (A)'
ws[f'{VOUT_COL}1'] = '输出电压 (V)'
ws[f'{IOUT_COL}1'] = '输出电流 (A)'
ws[f'{EFF_SYS_COL}1'] = '系统效率 (%)'
sws[f'{PLOSS_SYS_COL}1'] = '系统功耗 (W)'
excel_row_index = 2

# 初始化 VISA 资源管理器和仪器
rm = visa.ResourceManager()
print("可用的 VISA 资源:", rm.list_resources())

load = None
vin_dc = None
vcc_supply = None
daQ = None

try:
    # 打开与仪器的通信
    load       = rm.open_resource(LOAD_ADDRESS)
    vin_dc     = rm.open_resource(VIN_ADDRESS)
    vcc_supply = rm.open_resource(VCC_ADDRESS)
    daq        = rm.open_resource(DAQ_ADDRESS)

    # 打印仪器标识以供验证
    print("负载 ID:", load.query("*IDN?").strip())
    print("Vin 电源 ID:", vin_dc.query("*IDN?").strip())
    print("Vcc 电源 ID:", vcc_supply.query("*IDN?").strip())
    print("DAQ ID:", daq.query("*IDN?").strip())

    # 仪器基本设置
    load.write('CONF:REM ON')
    load.write('CHAN 3')
    load.write('MODE CCH')
    load.write('CURR:STAT:L1 0')

    # 获取测试条件
    print("\n请输入测试参数")
    if AC_MODE:
        # 交流模式额外输入
        AC_DUTY_CYCLE = float(input("占空比 (Duty Cycle) [%]: "))
        AC_WAVEFORM   = input("波形 (SINE/SQUARE/TRIANGLE): ")
        AC_FREQUENCY  = float(input("交流频率 [Hz]: "))
        print(f"AC 模式: 波形={AC_WAVEFORM}, 频率={AC_FREQUENCY}Hz, 占空比={AC_DUTY_CYCLE}%")
    else:
        vin_voltage       = float(input("输入电压 (Vin) [V]: "))
        vin_current_limit = float(input("输入电流限制 (Iin limit) [A]: "))
        vcc_voltage       = float(input("辅助电压 (Vcc) [V]: "))
        vcc_current_limit = float(input("辅助电流限制 (Icc limit) [A]: "))

    # TODO: 在此处根据 AC_MODE 调用相应的电源配置命令

    # 获取负载扫描参数
    print("\n请输入负载扫描参数")
    iout_current_min = int(input("最小输出电流 (Iout min) [A]: "))
    iout_current_max = int(input("最大输出电流 (Iout max) [A]: "))
    iout_current_step = int(input("输出电流步进 [A]: "))
    step_time         = float(input("每步停留时间 [s]: "))
    interval_time     = float(input("步进间隔时间 [s]: "))

    # 主测量循环
    print("\n开始效率测量扫描")
    for iout_setpoint in range(iout_current_min, iout_current_max + iout_current_step, iout_current_step):
        print(f"\n正在设置负载为 {iout_setpoint} A...")
        load.write(f"CURR:STAT:L1 {iout_setpoint}")
        load.write('LOAD ON')
        time.sleep(step_time)

        # 测量
        vin = float(daq.query('MEAS:VOLT:DC? 100, 1E-4, (@101)'))
        iin_sense = float(daq.query('MEAS:VOLT? 2, 1E-5, (@102)'))
        vout = float(daq.query('MEAS:VOLT? 100, 1E-4, (@103)'))
        iout_sense = float(daq.query('MEAS:VOLT? 1, 1E-5, (@104)'))

        load.write('CURR:STAT:L1 0')
        time.sleep(interval_time)

        # 计算
        iin = iin_sense / IIN_SHUNT_RESISTOR
        iout = iout_sense / IOUT_SHUNT_RESISTOR
        pin_sys   = vin * iin
        pout      = vout * iout
        ploss_sys = pin_sys - pout
        eff_sys   = (pout / pin_sys) * 100 if pin_sys > 0 else 0

        # 打印和记录
        print(f"输入电压={vin:.3f}V, 输入电流={iin:.3f}A, 输出电压={vout:.3f}V, 输出电流={iout:.3f}A")
        print(f"效率={eff_sys:.2f}%, 功率损耗={ploss_sys:.3f}W")

        ws[f'{VIN_COL}{excel_row_index}'] = vin
        ws[f'{IIN_COL}{excel_row_index}'] = iin
        ws[f'{VOUT_COL}{excel_row_index}'] = vout
        ws[f'{IOUT_COL}{excel_row_index}'] = iout
        ws[f'{EFF_SYS_COL}{excel_row_index}'] = eff_sys
        ws[f'{PLOSS_SYS_COL}{excel_row_index}'] = ploss_sys
        excel_row_index += 1
        efficiency_file.save(EXCEL_FILE_PATH)

    print("测量完成")

finally:
    if load:
        load.write('LOAD OFF')
        load.close()
        print("负载已关闭。")
    if vin_dc and not AC_MODE:
        vin_dc.write('CONF:OUTP OFF')
        vin_dc.close()
        print("Vin 电源已关闭。")
    if vcc_supply:
        vcc_supply.write('OUTP CH1,OFF')
        vcc_supply.close()
        print("Vcc 电源已关闭。")
    if daq:
        daq.close()
        print("DAQ 已关闭。")

    efficiency_file.save(EXCEL_FILE_PATH)
    print("Excel 已保存。")
