"""
电源转换器效率自动测量脚本

硬件:
1.  用于Vin的直流电源 (IT6833)
2.  用于VCC的直流电源 (Rigol DP832)
3.  可编程负载 (Chroma 6310A)
4.  数字万用表 (Agilent 34970A)
5.  安装了NI-VISA的电脑。

设置:
-   将所有仪器连接到电脑。
-   确保在脚本中指定了正确的VISA资源地址。
-   DAQ（数据采集单元）分别在通道101, 102, 103, 和 104上测量
    Vin, V_sense_iin, Vout, 和 V_sense_iout。
-   输入电流(iin)是根据一个采样电阻上的电压(V_sense_iin)计算得出的。
-   输出电流(iout)是根据一个采样电阻上的电压(V_sense_iout)计算得出的。

"""
import visa
import time
from openpyxl import Workbook

# 配置

# Excel文件配置

EXCEL_FILE_PATH = r'E:\AutomaticEfficiencyForPlusCurrent\test_data_chinese.xlsx'
VIN_COL = 'A'
IIN_COL = 'B'
VOUT_COL = 'C'
IOUT_COL = 'D'
EFF_SYS_COL = 'E'
PLOSS_SYS_COL = 'F'

# 测量配置
# 测量输入电流的采样电阻值
IIN_SHUNT_RESISTOR = 0.01
# 测量输出电流的采样电阻值
IOUT_SHUNT_RESISTOR = 0.001

# 设备visa地址
LOAD_ADDRESS = 'ASRL7::INSTR'
VIN_ADDRESS = 'USB0::0x1698::0x0837::002000002608::INSTR'
VCC_ADDRESS = 'USB0::0x1AB1::0x0E11::DP8C193003485::INSTR'
DAQ_ADDRESS = 'USB0::0x0957::0x2007::MY49029470::INSTR'


# 初始化

# 初始化Excel工作簿
try:
    efficiency_file = Workbook()
    print('成功创建新的Excel工作簿。')
except IOError:
    print('错误：创建Excel工作簿失败。')
    exit() # 如果文件无法创建则退出

# 创建并准备工作表
ws = efficiency_file.active
ws.title = "效率数据"

# 向工作表写入表头
ws[f'{VIN_COL}1'] = '输入电压 (V)'
ws[f'{IIN_COL}1'] = '输入电流 (A)'
ws[f'{VOUT_COL}1'] = '输出电压 (V)'
ws[f'{IOUT_COL}1'] = '输出电流 (A)'
ws[f'{EFF_SYS_COL}1'] = '系统效率 (%)'
ws[f'{PLOSS_SYS_COL}1'] = '系统功耗 (W)'
excel_row_index = 2 # 从第二行开始写入数据

# 初始化VISA资源管理器和仪器
rm = visa.ResourceManager()
print("可用的VISA资源:", rm.list_resources())

# 将仪器变量初始化为None
load = None
vin_dc = None
vcc_supply = None
daq = None

# 此代码块确保即使在发生错误时，所有仪器也能被安全关闭
try:
    # 打开与仪器的通信
    load = rm.open_resource(LOAD_ADDRESS)
    vin_dc = rm.open_resource(VIN_DC_SUPPLY_ADDRESS)
    vcc_supply = rm.open_resource(VCC_SUPPLY_ADDRESS)
    daq = rm.open_resource(DAQ_ADDRESS)

    # 打印仪器标识以供验证
    print("负载ID:", load.query("*IDN?").strip())
    print("Vin电源ID:", vin_dc.query("*IDN?").strip())
    print("Vcc电源ID:", vcc_supply.query("*IDN?").strip())
    print("DAQ ID:", daq.query("*IDN?").strip())


    # 仪器设置

    # 配置电子负载
    load.write('CONF:REM ON')  # 启用远程控制
    load.write('CHAN 3')       # 选择通道3，可更改
    load.write('MODE CCH')     # 设置为恒流高精度模式
    load.write('CURR:STAT:L1 0') # 设置初始电流为0A

    # 获取输入的测试条件
    print("\n请输入测试参数")
    vin_voltage = float(input("输入电压 (Vin) [V]: "))
    vin_current_limit = float(input("输入电流限制 (Iin limit) [A]: "))
    vcc_voltage = float(input("辅助电压 (Vcc) [V]: "))
    vcc_current_limit = float(input("辅助电流限制 (Icc limit) [A]: "))

    # 配置并启用Vin电源
    vin_dc.write(f"SOUR:VOLT {vin_voltage}")
    vin_dc.write(f"SOUR:CURR {vin_current_limit}")
    vin_dc.write('CONF:OUTP ON')
    print(f"Vin电源已设置为 {vin_voltage}V / {vin_current_limit}A 并已开启。")

    # 配置并启用VCC电源
    vcc_supply.write("INST CH1") # 选择通道1
    vcc_supply.write(f"VOLT {vcc_voltage}")
    vcc_supply.write(f"CURR {vcc_current_limit}")
    vcc_supply.write("OUTP CH1,ON")
    print(f"Vcc电源已设置为 {vcc_voltage}V / {vcc_current_limit}A 并已开启。")

    # 获取输入的负载扫描参数
    print("\n请输入负载扫描参数")
    iout_current_min = int(input("最小输出电流 (Iout min) [A]: "))
    iout_current_max = int(input("最大输出电流 (Iout max) [A]: "))
    iout_current_step = int(input("输出电流步进值 [A]: "))
    step_time = float(input("每一步的停留时间 [s]: "))
    interval_time = float(input("步进之间的间隔时间 [s]: "))


    # 主测量循环
    print("\n开始效率测量扫描")

    # 从最小电流循环到最大电流
    for iout_setpoint in range(iout_current_min, iout_current_max + iout_current_step, iout_current_step):
        print(f"\n正在设置负载为 {iout_setpoint} A...")

        # 设置负载电流并打开负载
        load.write(f"CURR:STAT:L1 {iout_setpoint}")
        load.write('LOAD ON')
        time.sleep(step_time) # 等待系统稳定

        # 使用DAQ进行测量
        # 通道101: 输入电压 (Vin)
        vin = float(daq.query('MEAS:VOLT:DC? 100, 1E-4, (@101)'))
        # 通道102: 输入电流采样电压
        iin_sense_voltage = float(daq.query('MEAS:VOLT? 2, 1E-5, (@102)'))
        # 通道103: 输出电压 (Vout)
        vout = float(daq.query('MEAS:VOLT? 100, 1E-4, (@103)'))
        # 通道104: 输出电流采样电压
        iout_sense_voltage = float(daq.query('MEAS:VOLT? 1, 1E-5, (@104)'))

        # 在间隔期间将负载电流设置为0A以冷却
        load.write('CURR:STAT:L1 0')
        time.sleep(interval_time)

        # 计算最终值
        iin = iin_sense_voltage / IIN_SHUNT_RESISTOR
        iout = iout_sense_voltage / IOUT_SHUNT_RESISTOR

        # 避免输入功率为零时出现除零错误
        if vin > 0 and iin > 0:
            pin_sys = vin * iin
            pout = vout * iout
            ploss_sys = pin_sys - pout
            eff_sys = (pout / pin_sys) * 100
        else:
            pin_sys = 0
            pout = 0
            ploss_sys = 0
            eff_sys = 0

        # 将结果打印到控制台
        print(f"输入电压={vin:.3f}V, 输入电流={iin:.3f}A, 输出电压={vout:.3f}V, 输出电流={iout:.3f}A")
        print(f"效率={eff_sys:.2f}%, 功率损耗={ploss_sys:.3f}W")

        # 将数据记录到Excel
        ws[f'{VIN_COL}{excel_row_index}'] = vin
        ws[f'{IIN_COL}{excel_row_index}'] = iin
        ws[f'{VOUT_COL}{excel_row_index}'] = vout
        ws[f'{IOUT_COL}{excel_row_index}'] = iout
        ws[f'{EFF_SYS_COL}{excel_row_index}'] = eff_sys
        ws[f'{PLOSS_SYS_COL}{excel_row_index}'] = ploss_sys
        excel_row_index += 1 # 移动到下一行

        efficiency_file.save(EXCEL_FILE_PATH)

    print("\n--- 测量扫描完成 ---")


# 清理
finally:
    print("\n--- 正在清理并关闭仪器 ---")
    if load:
        load.write('LOAD OFF') # 关闭电子负载
        load.close()
        print("负载已关闭。")
    if vin_dc:
        vin_dc.write('CONF:OUTP OFF') # 关闭主电源输出
        vin_dc.close()
        print("Vin电源已关闭。")
    if vcc_supply:
        vcc_supply.write("OUTP CH1,OFF") # 关闭VCC电源输出
        vcc_supply.close()
        print("Vcc电源已关闭。")
    if daq:
        daq.close()

    # 最后一次保存Excel文件
    efficiency_file.save(EXCEL_FILE_PATH)
